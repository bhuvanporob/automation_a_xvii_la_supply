import streamlit as st
import pandas as pd
import pymysql
from sqlalchemy import create_engine
import plotly.express as px
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# Database connection settings
DB_USER = "root"
DB_PASSWORD = "your_password"
DB_HOST = "localhost"
DB_PORT = "3306"
DB_NAME = "mtb_supply"

# Create a database connection
@st.cache_resource(ttl=600)  # Refreshes every 10 minutes
def get_db_connection():
    return create_engine(f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

# Function to load and process data based on the selected page
def load_and_process_data(table_name, start_date, end_date, lot_no, page):
    engine = get_db_connection()
    query = f"""
        SELECT * FROM {table_name}
        WHERE Test_date_time BETWEEN '{start_date}' AND '{end_date}'
        AND Lot IN ({lot_no})
    """
    df = pd.read_sql(query, con=engine)
    
    # Data Cleaning
    df = df[df['Test_status'].notnull()]
    df = df[~df['Chip_serial_no'].str[0].str.isdigit()]
    df = df[df['Chip_serial_no'].str[1].str.isdigit()]

    # Define pivoting logic
    index_columns = {
        "Lot Performance": ['Lot'],
        "Chip series": ['Chip_serial_no'],
        "Lot Chip series": ['Lot', 'Chip_serial_no'],
        "Lot chip batch chip series": ['Lot', 'Chip_batchno', 'Chip_serial_no'],
        "Detailed Data": ['Lab_name', 'Truelab_id', 'Lot', 'Chip_batchno', 'Chip_serial_no']
    }
    
    if page in index_columns:
        pivot = df.pivot_table(index=index_columns[page], values='Patient_id', columns='Test_status', aggfunc='count', margins=True)
        pivotdf = pd.DataFrame(pivot.to_records())
        
        columns_to_fill = ['Error-1', 'Error-1A', 'Error-2', 'Error-3', 'Error-4', 'Error-5', 'Invalid', 'Valid', 'All']
        available_columns = set(pivotdf.columns) & set(columns_to_fill)
        pivotdf[list(available_columns)] = pivotdf[list(available_columns)].fillna(0)
        
        if "Invalid" in pivotdf.columns and "All" in pivotdf.columns:
            pivotdf["INV_per"] = round(pivotdf["Invalid"] / pivotdf["All"] * 100, 2)
        else:
            pivotdf["INV_per"] = 0
    else:
        pivotdf = df
    
    return df, pivotdf

# Function for conditional formatting in Excel
def apply_excel_formatting(writer, sheet_name, df):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # Define colors and styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    alternate_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    green_fill = PatternFill(start_color="B3E7B5", end_color="B3E7B5", fill_type="solid")
    red_fill = PatternFill(start_color="FDB1B1", end_color="FDB1B1", fill_type="solid")
    green_font = Font(color="006100")
    red_font = Font(color="9C0006")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Apply header formatting
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.border = thin_border

    # Apply alternating row colors
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        fill = white_fill if row_idx % 2 == 0 else alternate_fill
        for cell in row:
            cell.fill = fill
            cell.border = thin_border

    # Apply formatting to the 'INV_per' column
    inv_col_idx = None
    for col_idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == "INV_per":
            inv_col_idx = col_idx
            break

    if inv_col_idx:
        for row_idx, value in enumerate(df["INV_per"], start=2):
            cell = worksheet.cell(row=row_idx, column=inv_col_idx)
            if isinstance(value, (int, float)):
                if value > 7:
                    cell.fill = red_fill
                    cell.font = red_font
                else:
                    cell.fill = green_fill
                    cell.font = green_font
            cell.border = thin_border

# Sidebar
st.sidebar.title("Filters")
start_date = st.sidebar.date_input("Start Date")
end_date = st.sidebar.date_input("End Date")
category = st.sidebar.selectbox("Select Category", ["17 Lakhs", "16 Lakhs Tranche 1", "16 Lakhs Tranche 2"])

lot_mapping = {
    "17 Lakhs": "'076TB', '077TB', '078TB', '079TB', '080TB', '081TB', '082TB', '083TB', '084TB', '085TB', '086TB', '087TB', '088TB', '089TB', '090TB', '091TB', '092TB', '093TB', '094TB', '095TB', '096TB', '097TB', '098TB', '099TB', '100TB', '101TB', '102TB', '103TB', '104TB', '105TB', '106TB', '107TB', '108TB', '109TB', '110TB', '111TB', '112TB', '113TB', '114TB', '115TB', '116TB', '117TB', '118TB', '119TB', 'TB260', 'TB261', 'TB262', 'TB263', 'TB267', 'TB268', 'TB269', 'TB270', 'TB271', 'TB272', 'TB273'",
    "16 Lakhs Tranche 1": "'076TB', '077TB', '078TB'",
    "16 Lakhs Tranche 2": "'120TB', '121TB', '122TB', '123TB', '124TB', '125TB', '126TB', '127TB', '128TB', '129TB', '130TB', '131TB', '132TB', '133TB', '134TB', '135TB', '136TB', '137TB', '138TB', '139TB', '140TB'"
}
lot_no = lot_mapping.get(category, "''")

pages = {
    "Lot Performance": "all_test",
    "Chip series": "all_test",
    "Lot Chip series": "all_test",
    "Lot chip batch chip series": "all_test",
    "Detailed Data": "all_test"
}

selected_page = st.sidebar.radio("Choose a Page", list(pages.keys()))

table_name = pages[selected_page]
data, data_pivot = load_and_process_data(table_name, start_date, end_date, lot_no, selected_page)

st.subheader(f"📊 Processed Data for {selected_page}")
st.dataframe(data_pivot)

if not data_pivot.empty and "INV_per" in data_pivot.columns:
    data_pivot = data_pivot.reset_index()
    fig = px.line(
        data_pivot, x=data_pivot.columns[0], y="INV_per",
        title=f"INV% Trend for {selected_page}",
        labels={data_pivot.columns[0]: "Category", "INV_per": "Invalid Percentage (%)"},
        markers=True
    )
    st.plotly_chart(fig, use_container_width=True)

# Function to generate and download Excel
def generate_excel():
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for page, table in pages.items():
            _, df = load_and_process_data(table, start_date, end_date, lot_no, page)
            df.to_excel(writer, sheet_name=page, index=False)
            apply_excel_formatting(writer, page, df)
    output.seek(0)
    return output

st.sidebar.subheader("Download Data")
if st.sidebar.button("Download All Pages as Excel"):
    excel_file = generate_excel()
    st.sidebar.download_button(
        label="📥 Download Excel File",
        data=excel_file,
        file_name="data_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
